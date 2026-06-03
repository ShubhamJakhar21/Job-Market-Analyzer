"""
analyzer.py  —  Cleans job data, analyzes skills, creates charts + Excel report
"""

import os
import re
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from collections import Counter


# ── Colors used in charts ──────────────────────────────────────────────────
CHART_COLORS = [
    "#4C9BE8", "#5CB8B2", "#F4845F", "#A78BFA",
    "#FBBF24", "#34D399", "#F472B6", "#60A5FA",
    "#FB923C", "#A3E635", "#2DD4BF", "#818CF8",
    "#E879F9", "#FCD34D", "#6EE7B7", "#93C5FD",
    "#FCA5A5", "#86EFAC", "#C4B5FD", "#FDE68A",
]


def load_data(csv_path: str = "jobs.csv") -> pd.DataFrame:
    """Load scraped jobs CSV into a DataFrame."""
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"❌ '{csv_path}' not found. Run main.py first to scrape data.")
    df = pd.read_csv(csv_path)
    print(f"📂 Loaded {len(df)} rows from {csv_path}")
    return df


def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Remove duplicates, fix whitespace, drop useless rows."""
    before = len(df)
    df = df.drop_duplicates(subset=["title", "company"])
    df = df[df["title"] != "N/A"]
    df = df.dropna(subset=["title"])

    # Clean text columns
    for col in ["title", "company", "location", "skills"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    print(f"🧹 Cleaned: {before} → {len(df)} rows (removed {before - len(df)} duplicates/blanks)")
    return df.reset_index(drop=True)


def extract_all_skills(df: pd.DataFrame) -> pd.Series:
    """
    Splits the comma-separated 'skills' column into individual skill tokens,
    then counts how many jobs demand each skill.
    """
    skill_counter = Counter()

    for raw in df["skills"].dropna():
        if raw == "N/A":
            continue
        parts = re.split(r"[,\|/\n]", raw)   # split on comma, pipe, slash, newline
        for part in parts:
            skill = part.strip().title()       # "python" → "Python"
            if skill and len(skill) >= 2:
                skill_counter[skill] += 1

    return pd.Series(skill_counter).sort_values(ascending=False)


def get_top_locations(df: pd.DataFrame, top_n: int = 10) -> pd.Series:
    """Count jobs per city (first city mentioned in the location field)."""
    cities = (
        df["location"]
        .dropna()
        .str.split(",")
        .str[0]
        .str.strip()
        .str.title()
    )
    cities = cities[cities != "N/A"]
    return cities.value_counts().head(top_n)


def get_top_companies(df: pd.DataFrame, top_n: int = 10) -> pd.Series:
    """Count job postings per company."""
    companies = df["company"].dropna()
    companies = companies[companies != "N/A"]
    return companies.value_counts().head(top_n)


# ── Chart helpers ───────────────────────────────────────────────────────────

def _bar_chart(series: pd.Series, title: str, xlabel: str,
               filepath: str, color_list: list = None):
    """Generic horizontal bar chart. Saves to filepath."""
    fig, ax = plt.subplots(figsize=(12, 6))
    colors = (color_list or CHART_COLORS)[:len(series)]

    bars = ax.barh(series.index[::-1], series.values[::-1], color=colors[::-1])
    ax.set_title(title, fontsize=16, fontweight="bold", pad=16)
    ax.set_xlabel(xlabel, fontsize=12)
    ax.xaxis.set_major_locator(mticker.MaxNLocator(integer=True))

    # Value labels on bars
    for bar in bars:
        width = bar.get_width()
        ax.text(width + 0.2, bar.get_y() + bar.get_height() / 2,
                f"{int(width)}", va="center", fontsize=9)

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    plt.tight_layout()
    plt.savefig(filepath, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"   📊 Saved chart: {filepath}")


def create_charts(top_skills: pd.Series,
                  top_locations: pd.Series,
                  top_companies: pd.Series,
                  output_dir: str = "."):
    """Generate all three charts and return file paths."""
    os.makedirs(output_dir, exist_ok=True)

    skills_path    = os.path.join(output_dir, "chart_skills.png")
    locations_path = os.path.join(output_dir, "chart_locations.png")
    companies_path = os.path.join(output_dir, "chart_companies.png")

    _bar_chart(top_skills.head(20),    "Top 20 In-Demand Skills",          "Number of Job Postings", skills_path)
    _bar_chart(top_locations.head(10), "Top 10 Hiring Locations",          "Number of Job Postings", locations_path, CHART_COLORS[4:])
    _bar_chart(top_companies.head(10), "Top 10 Companies Hiring",          "Number of Job Postings", companies_path, CHART_COLORS[8:])

    return skills_path, locations_path, companies_path


# ── Excel report ────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F3864")   # Dark navy
ALT_FILL      = PatternFill("solid", start_color="EFF3FB")   # Light blue
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
BODY_FONT     = Font(name="Arial", size=10)
TITLE_FONT    = Font(bold=True, name="Arial", size=14, color="1F3864")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7E8"),
    right=Side(style="thin", color="D0D7E8"),
    top=Side(style="thin", color="D0D7E8"),
    bottom=Side(style="thin", color="D0D7E8"),
)


def _write_sheet_header(ws, title: str, columns: list[str]):
    """Write a title row then a styled header row."""
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 30

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 22


def _write_data_rows(ws, data: list[list], start_row: int = 3):
    """Write alternating-row data with border."""
    for row_idx, row_data in enumerate(data, start=start_row):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = LEFT
            cell.border = THIN_BORDER


def _auto_col_width(ws, min_w=10, max_w=45):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def create_excel_report(df: pd.DataFrame,
                        top_skills: pd.Series,
                        top_locations: pd.Series,
                        top_companies: pd.Series,
                        chart_paths: tuple,
                        output_path: str = "job_market_report.xlsx"):
    """
    Builds a nicely formatted Excel workbook with four sheets:
      1. Summary        — key numbers at a glance
      2. All Jobs       — full scraped data
      3. Top Skills     — skill frequency table
      4. Charts         — the three bar charts embedded
    """
    print("\n📝 Creating Excel report…")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── Sheet 1: Summary ────────────────────────────────────────────────
        summary_data = {
            "Metric": [
                "Total Jobs Scraped",
                "Unique Companies",
                "Unique Locations",
                "Top Skill",
                "Top Location",
                "Top Company",
            ],
            "Value": [
                len(df),
                df["company"].nunique(),
                df["location"].nunique(),
                top_skills.index[0] if len(top_skills) > 0 else "N/A",
                top_locations.index[0] if len(top_locations) > 0 else "N/A",
                top_companies.index[0] if len(top_companies) > 0 else "N/A",
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False, startrow=1)

        # ── Sheet 2: All Jobs ───────────────────────────────────────────────
        df.to_excel(writer, sheet_name="All Jobs", index=False, startrow=1)

        # ── Sheet 3: Top Skills ─────────────────────────────────────────────
        skill_df = top_skills.reset_index()
        skill_df.columns = ["Skill", "Job Count"]
        skill_df["Rank"] = range(1, len(skill_df) + 1)
        skill_df = skill_df[["Rank", "Skill", "Job Count"]]
        skill_df.to_excel(writer, sheet_name="Top Skills", index=False, startrow=1)

        # ── Sheet 4: Top Locations ──────────────────────────────────────────
        loc_df = top_locations.reset_index()
        loc_df.columns = ["City", "Job Count"]
        loc_df.to_excel(writer, sheet_name="Top Locations", index=False, startrow=1)

        # ── Sheet 5: Charts (placeholder) ──────────────────────────────────
        pd.DataFrame().to_excel(writer, sheet_name="Charts", index=False)

    # ── Apply formatting with openpyxl ──────────────────────────────────────
    wb = load_workbook(output_path)

    # -- Summary sheet --
    ws = wb["Summary"]
    _write_sheet_header(ws, "📋 Job Market Analysis — Summary", ["Metric", "Value"])
    _auto_col_width(ws)

    # -- All Jobs sheet --
    ws = wb["All Jobs"]
    cols = list(df.columns)
    _write_sheet_header(ws, "📄 All Scraped Job Listings", cols)
    ws.freeze_panes = "A3"
    _auto_col_width(ws)

    # -- Top Skills sheet --
    ws = wb["Top Skills"]
    _write_sheet_header(ws, "🔧 Most In-Demand Skills", ["Rank", "Skill", "Job Count"])
    _auto_col_width(ws)

    # -- Top Locations sheet --
    ws = wb["Top Locations"]
    _write_sheet_header(ws, "📍 Top Hiring Locations", ["City", "Job Count"])
    _auto_col_width(ws)

    # -- Charts sheet: embed PNG images --
    ws_charts = wb["Charts"]
    ws_charts["A1"] = "Job Market Charts"
    ws_charts["A1"].font = TITLE_FONT

    chart_titles = ["Top 20 Skills", "Top Locations", "Top Companies Hiring"]
    row_positions = [3, 25, 47]
    col_positions = ["A", "A", "A"]

    for path, row, col, title in zip(chart_paths, row_positions, col_positions, chart_titles):
        if os.path.exists(path):
            label_cell = ws_charts[f"{col}{row - 1}"]
            label_cell.value = title
            label_cell.font = Font(bold=True, name="Arial", size=12, color="1F3864")
            img = XLImage(path)
            img.width  = 780
            img.height = 390
            ws_charts.add_image(img, f"{col}{row}")

    wb.save(output_path)
    print(f"✅ Excel report saved: {output_path}")


def run_analysis(csv_path: str = "jobs.csv",
                 output_dir: str = "output",
                 top_n_skills: int = 20):
    """Full pipeline: load → clean → analyze → charts → Excel."""
    os.makedirs(output_dir, exist_ok=True)

    print("\n" + "=" * 55)
    print("  JOB MARKET ANALYZER — Analysis Started")
    print("=" * 55)

    df = load_data(csv_path)
    df = clean_data(df)

    print("\n📊 Analyzing data…")
    top_skills    = extract_all_skills(df).head(top_n_skills)
    top_locations = get_top_locations(df)
    top_companies = get_top_companies(df)

    print(f"\n🏆 Top 5 skills: {', '.join(top_skills.head(5).index.tolist())}")
    print(f"📍 Top 3 cities: {', '.join(top_locations.head(3).index.tolist())}")

    print("\n📈 Generating charts…")
    chart_paths = create_charts(top_skills, top_locations, top_companies, output_dir)

    excel_path = os.path.join(output_dir, "job_market_report.xlsx")
    create_excel_report(df, top_skills, top_locations, top_companies, chart_paths, excel_path)

    print("\n" + "=" * 55)
    print("  ✅ Analysis Complete!")
    print(f"  📁 All files saved in: ./{output_dir}/")
    print("=" * 55)

    return df, top_skills, top_locations, top_companies


if __name__ == "__main__":
    run_analysis()
